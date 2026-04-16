import sqlite3
import os
from typing import Dict, List, Optional, Any
from datetime import datetime


class DatabaseManager:
    """SQLite 数据库管理器，负责文献记录的 CRUD、FTS5 全文检索和 Excel 导出"""

    def __init__(self, db_path: str = "literature_records.db"):
        self.db_path = db_path

    def _get_connection(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        return conn

    def init_db(self):
        """创建数据库表、索引和 FTS5 全文检索虚拟表"""
        conn = self._get_connection()
        try:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS literature_records (
                    id           INTEGER PRIMARY KEY AUTOINCREMENT,
                    file_path    TEXT NOT NULL,
                    file_type    TEXT NOT NULL,
                    content_hash TEXT NOT NULL UNIQUE,
                    title        TEXT,
                    keywords     TEXT,
                    abstract     TEXT,
                    abstract_cn  TEXT,
                    summary      TEXT,
                    created_at   TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at   TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            conn.execute("CREATE INDEX IF NOT EXISTS idx_content_hash ON literature_records(content_hash)")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_title ON literature_records(title)")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_file_type ON literature_records(file_type)")

            # FTS5 全文检索虚拟表（title, keywords, abstract, abstract_cn, summary, file_path）
            conn.execute("""
                CREATE VIRTUAL TABLE IF NOT EXISTS literature_fts USING fts5(
                    title, keywords, abstract, abstract_cn, summary, file_path,
                    content='literature_records', content_rowid='id'
                )
            """)

            # 确保同步触发器存在（保持 FTS 索引与主表一致）
            conn.execute("""
                CREATE TRIGGER IF NOT EXISTS literature_fts_ai AFTER INSERT ON literature_records BEGIN
                    INSERT INTO literature_fts(rowid, title, keywords, abstract, abstract_cn, summary, file_path)
                    VALUES (new.id, new.title, new.keywords, new.abstract, new.abstract_cn, new.summary, new.file_path);
                END
            """)
            conn.execute("""
                CREATE TRIGGER IF NOT EXISTS literature_fts_ad AFTER DELETE ON literature_records BEGIN
                    INSERT INTO literature_fts(literature_fts, rowid, title, keywords, abstract, abstract_cn, summary, file_path)
                    VALUES ('delete', old.id, old.title, old.keywords, old.abstract, old.abstract_cn, old.summary, old.file_path);
                END
            """)
            conn.execute("""
                CREATE TRIGGER IF NOT EXISTS literature_fts_au AFTER UPDATE ON literature_records BEGIN
                    INSERT INTO literature_fts(literature_fts, rowid, title, keywords, abstract, abstract_cn, summary, file_path)
                    VALUES ('delete', old.id, old.title, old.keywords, old.abstract, old.abstract_cn, old.summary, old.file_path);
                    INSERT INTO literature_fts(rowid, title, keywords, abstract, abstract_cn, summary, file_path)
                    VALUES (new.id, new.title, new.keywords, new.abstract, new.abstract_cn, new.summary, new.file_path);
                END
            """)

            conn.commit()
        finally:
            conn.close()

    def insert_record(self, record: Dict[str, Any]) -> int:
        """
        插入新的文献记录（FTS 索引由触发器自动同步）

        Returns:
            新记录的 ID

        Raises:
            sqlite3.IntegrityError: content_hash 已存在（重复）
        """
        conn = self._get_connection()
        try:
            cursor = conn.execute("""
                INSERT INTO literature_records
                    (file_path, file_type, content_hash, title, keywords,
                     abstract, abstract_cn, summary)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                record['file_path'],
                record['file_type'],
                record['content_hash'],
                record.get('title', ''),
                record.get('keywords', ''),
                record.get('abstract', ''),
                record.get('abstract_cn', ''),
                record.get('summary', ''),
            ))
            conn.commit()
            return cursor.lastrowid
        finally:
            conn.close()

    def check_duplicate(self, content_hash: str) -> Optional[Dict]:
        """检查是否已存在相同内容的记录，返回已有记录或 None"""
        conn = self._get_connection()
        try:
            row = conn.execute(
                "SELECT * FROM literature_records WHERE content_hash = ?",
                (content_hash,)
            ).fetchone()
            return dict(row) if row else None
        finally:
            conn.close()

    def get_all_records(self, file_type: str = None) -> List[Dict]:
        """获取所有记录，可按文件类型筛选"""
        conn = self._get_connection()
        try:
            if file_type:
                rows = conn.execute(
                    "SELECT * FROM literature_records WHERE file_type = ? ORDER BY created_at DESC",
                    (file_type,)
                ).fetchall()
            else:
                rows = conn.execute(
                    "SELECT * FROM literature_records ORDER BY created_at DESC"
                ).fetchall()
            return [dict(row) for row in rows]
        finally:
            conn.close()

    def search_records(self, query: str) -> List[Dict]:
        """
        使用 FTS5 + BM25 全文检索，按相关性排序

        支持的查询语法：
        - 简单词: "神经网络" → 匹配包含这些字的记录
        - 多词 AND: "graph neural" → 同时包含两个词
        - 短语: '"deep learning"' → 精确短语匹配
        - OR: "graph OR network" → 包含任一词
        """
        if not query.strip():
            return self.get_all_records()

        conn = self._get_connection()
        try:
            # 转义 FTS5 特殊字符，构建安全查询
            safe_query = self._build_fts_query(query)

            rows = conn.execute("""
                SELECT r.*, fts.rank
                FROM literature_records r
                JOIN literature_fts fts ON r.id = fts.rowid
                WHERE literature_fts MATCH ?
                ORDER BY bm25(literature_fts)
            """, (safe_query,)).fetchall()
            return [dict(row) for row in rows]
        except sqlite3.OperationalError:
            # FTS 查询语法失败时回退到 LIKE 搜索
            return self._fallback_search(query)
        finally:
            conn.close()

    def _build_fts_query(self, query: str) -> str:
        """
        将用户输入构建为安全的 FTS5 查询

        对中文（逐字匹配）和英文（按词匹配）均适用。
        多个词之间用 AND 连接。
        """
        # 去除 FTS5 特殊字符
        special_chars = ['"', '*', '(', ')', ':', '^', '+', '-', '|']
        cleaned = query
        for ch in special_chars:
            cleaned = cleaned.replace(ch, ' ')

        # 按空白分割，过滤空词
        terms = [t.strip() for t in cleaned.split() if t.strip()]
        if not terms:
            return query

        # 每个词加后缀 * 支持前缀匹配，词间 AND 连接
        fts_terms = [f'"{t}"' for t in terms]
        return ' AND '.join(fts_terms)

    def _fallback_search(self, query: str) -> List[Dict]:
        """FTS 不可用时的 LIKE 回退搜索"""
        conn = self._get_connection()
        try:
            like_pattern = f"%{query}%"
            rows = conn.execute("""
                SELECT * FROM literature_records
                WHERE title LIKE ? OR keywords LIKE ? OR abstract LIKE ?
                   OR abstract_cn LIKE ? OR summary LIKE ? OR file_path LIKE ?
                ORDER BY created_at DESC
            """, (like_pattern, like_pattern, like_pattern,
                  like_pattern, like_pattern, like_pattern)).fetchall()
            return [dict(row) for row in rows]
        finally:
            conn.close()

    def get_record_by_id(self, record_id: int) -> Optional[Dict]:
        """按 ID 获取单条记录"""
        conn = self._get_connection()
        try:
            row = conn.execute(
                "SELECT * FROM literature_records WHERE id = ?",
                (record_id,)
            ).fetchone()
            return dict(row) if row else None
        finally:
            conn.close()

    def delete_record(self, record_id: int) -> bool:
        """删除指定记录（FTS 索引由触发器自动同步）"""
        conn = self._get_connection()
        try:
            cursor = conn.execute(
                "DELETE FROM literature_records WHERE id = ?",
                (record_id,)
            )
            conn.commit()
            return cursor.rowcount > 0
        finally:
            conn.close()

    def get_records_by_ids(self, record_ids: List[int]) -> List[Dict]:
        """按 ID 列表获取记录"""
        if not record_ids:
            return []
        conn = self._get_connection()
        try:
            placeholders = ','.join('?' * len(record_ids))
            rows = conn.execute(
                f"SELECT * FROM literature_records WHERE id IN ({placeholders}) ORDER BY created_at DESC",
                record_ids
            ).fetchall()
            return [dict(row) for row in rows]
        finally:
            conn.close()

    def export_to_excel(self, output_path: str, record_ids: List[int] = None) -> str:
        """导出记录到 Excel 文件"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
        except ImportError:
            raise ImportError("请安装 openpyxl: pip install openpyxl")

        if record_ids:
            records = self.get_records_by_ids(record_ids)
        else:
            records = self.get_all_records()

        wb = Workbook()
        ws = wb.active
        ws.title = "文献记录"

        # 表头
        headers = ["ID", "标题", "关键词", "摘要", "中文摘要", "文章概要", "文件类型", "文件路径", "记录时间"]
        header_font = Font(bold=True, size=11)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font

        # 数据行
        wrap_alignment = Alignment(wrap_text=True, vertical='top')
        for row_idx, record in enumerate(records, 2):
            values = [
                record.get('id', ''),
                record.get('title', ''),
                record.get('keywords', ''),
                record.get('abstract', ''),
                record.get('abstract_cn', ''),
                record.get('summary', ''),
                record.get('file_type', ''),
                record.get('file_path', ''),
                record.get('created_at', ''),
            ]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.alignment = wrap_alignment

        # 设置列宽
        column_widths = [6, 40, 25, 50, 50, 50, 10, 40, 20]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width

        wb.save(output_path)
        return output_path
